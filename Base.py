# -*- coding: utf-8 -*-
"""
Created on Tue Apr 11 07:55:55 2017

@author: 
"""

"""
Any line win or Expanded Pay will advance the Speed Spins Active Reels two reels to the right and may increase the Speed Spins
Multiplier up to 5x.
This is the reason for the 9 different reels as well as the multiplier and the nested calculations.

"""

import os
import pandas as pd
import numpy as np
import random as rd


from openpyxl import load_workbook
from openpyxl.utils import get_column_interval
import re

def load_workbook_range(range_string, ws):
    start, end = range_string.split(':')
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[start:end]:
        data_rows.append([cell.value for cell in row])

    return pd.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))




########################
directory = os.getcwd()
#filename has been removed to 
filename = 'SlotGame.xlsx'
filename = os.path.join(directory,filename)

wb = load_workbook(filename, data_only = True)

BaseReels = load_workbook_range('C11:K73', wb.get_sheet_by_name('Reels'))
JackpotReel = load_workbook_range('L11:L106',wb.get_sheet_by_name('Reels'))

Jackpot = load_workbook_range('D17:E19',wb.get_sheet_by_name('Jackpot'))



Bet = 35
NumLines = 9
Lines = pd.DataFrame([[1,1,1],
                      [0,0,0],
                      [2,2,2],
                      [0,1,2],
                      [2,1,0],
                      [1,0,1],
                      [1,2,1],
                      [0,1,0],
                      [2,1,2]])
    
WindowSize = [3]*9
Wild = 'WILD'    
Scatter = 'BONUS'
BasePay = {'WILD':0,'H1':100,'H2':50,'H3':25,'M1':15,'M2':15,'M3':15,
           'L1':5,'L2':5,'L3':5,'L4':5,
           'ANYH':10,'ANYM':5}
ExpandedPay = {'H1':2000,'H2':1000,'H3':750,'M1':500,'M2':350,'M3':250,
               'L1':175,'L2':175,'L3':175,'L4':175,
               'ANYH':250,'ANYM':200,'ANYL':50}

AnyH = ['H1','H2','H3',Wild]
AnyM = ['M1','M2','M3',Wild]
AnyL = ['L1','L2','L3',Wild]


def GetReelLength(ReelSet):
    L1 = ReelSet.iloc[:,0].count()
    L2 = ReelSet.iloc[:,1].count()
    L3 = ReelSet.iloc[:,2].count()
    L4 = ReelSet.iloc[:,3].count()
    L5 = ReelSet.iloc[:,4].count()
    L6 = ReelSet.iloc[:,5].count()
    L7 = ReelSet.iloc[:,6].count()
    L8 = ReelSet.iloc[:,7].count()
    L9 = ReelSet.iloc[:,8].count()
    
    return([L1,L2,L3,L4,L5,L6,L7,L8,L9])

BaseReelLength = GetReelLength(BaseReels)

def GetStop(ReelLength):
    Stop = [0]*len(ReelLength)
    for i in range(len(ReelLength)):
        Stop[i] = rd.randint(0,ReelLength[i]-1)
    #return([18, 49, 41, 46, 25, 48, 46, 3, 19])    
    return(Stop)

#need 2 different GetWindows because of the different reel sizes    
def GetFullWindow(ReelLength):
    Stop = GetStop(ReelLength)
    #print(Stop)
    #Stop = [18,27,30,31,31]    
    
    Window = pd.DataFrame(np.nan, index = [0,1,2],
                          columns = ['R1','R2','R3','R4','R5','R6','R7','R8','R9'])   
    
    #creates a window of the visisble symbols. this needs to be changed so that offset can
    #be manually set. this calc has top spot as 0 offset
    for i in range(9):
        for x in range(WindowSize[i]):
            Window.iloc[x,i] = BaseReels.iloc[(Stop[i]+x) % ReelLength[i],i]
        
    return(Window)


def GetScatterCount(Window):
    Result = Window.apply(pd.value_counts).fillna(0)
    
    try:
        ScatterCount = Result.sum(axis = 1)[Scatter]
        if  ScatterCount >= 3:
            return(ScatterCount)
        else:
            return(0)
    except:
        return(0)


def GetLine(Window, LineNum):
    Line = ['aaple']*3
    LineOffsets = Lines.iloc[LineNum,:]
    
    for i in range(3):
        Line[i] = Window.iloc[:,i][LineOffsets.iloc[i]]
        
    return(Line)
    
def GetLineWin(Line,Type):
    WildCount = 0
    SymbolCount = 0
    HCount = 0
    MCount = 0
    Symbol = 'apple'
    
    for i in range(3):
        if Symbol == 'apple':
            if Line[i] == Wild:
                WildCount += 1
                SymbolCount += 1
                HCount += 1
                MCount += 1
            elif Line[i] != Scatter:
                Symbol = Line[i]
                SymbolCount += 1
                
                if Line[i] in AnyH:
                    HCount += 1
                elif Line[i] in AnyM:
                    MCount += 1
            else:
                break
        else:
            if Line[i] == Wild:
                WildCount += 1
                SymbolCount += 1
                HCount += 1
                MCount += 1
            elif Line[i] == Symbol:
                SymbolCount += 1
                
                if Line[i] in AnyH:
                    HCount += 1
                elif Line[i] in AnyM:
                    MCount += 1
            elif Line[i] in AnyH:
                HCount += 1
            elif Line[i] in AnyM:
                MCount += 1
                
            else:
                break
        
        
        
        
        
        
        
    return(max(GetSymbolPay(Symbol,SymbolCount,Type),
               GetSymbolPay(Wild,WildCount,Type),
                GetSymbolPay('ANYH',HCount,Type),
                GetSymbolPay('ANYM',MCount,Type))) 
    

def GetSymbolPay(Symbol,Count,Type):
    try:
        if Type == 'Base':
            if Count == 3:
                return(BasePay[Symbol])
            else:
                return(0)
    except:
        return(0)

def PlayWindow(Window,Mult):
    WindowPay = 0
    
    WindowPay += GetExpPay(Window)
    
    if WindowPay == 0:
        for i in range(len(Lines)):
            Line = GetLine(Window,i)
            WindowPay += GetLineWin(Line,'Base')
            
    return(WindowPay*Mult)


def GetJackpotPay():
    #Win = 'JP'
    #x = rd.randint(0,len(JackpotReel)-1)
    #print(x)
    Win = JackpotReel.iloc[rd.randint(0,len(JackpotReel)-1),0]
    #print(Win)
    
    if type(Win) == 'int':
        return(Win)
    elif Win == 'BLANK':
        return(0)
    else:
        JP = [0]
        
        for i in range(3):
            JP.append(Jackpot.iloc[i,1]+JP[i])
            
        stop = rd.randint(0,JP[3]-1)
        
        for i in range(4):
            if stop < JP[i]:
                return(Jackpot.iloc[i-1,0])
          
def GetExpPay(Window):
    SingleSymbols = ['H1','H2','H3','M1','M2','M3',
                     'L1','L2','L3','L4',AnyH,AnyM,AnyL]    
    
    for sym in SingleSymbols:
        SymCount = 0 
        for row in range(3):
            for col in range(3):
                if Window.iloc[row,col] in sym:
                    SymCount += 1
                else:
                    break         
                
                

        if SymCount == 9:
            if sym == AnyH:
                return(ExpandedPay['ANYH'])
            elif sym == AnyM:
                return(ExpandedPay['ANYM'])
            elif sym == AnyL:
                return(ExpandedPay['ANYL'])
            else:
                return(ExpandedPay[sym])
            
        
    return(0)
    
def PlayBaseOnce():
    FullWindow = GetFullWindow(BaseReelLength)
    
    #FullWindow = Window
    ActiveWindow = FullWindow.iloc[:,0:(2+1)]
    #print(ActiveWindow)
    WindowPay = 0
    Mult = 1
    
    
    WindowPay += PlayWindow(ActiveWindow,Mult)
    #print(WindowPay)                           
                               
    if WindowPay > 0 :
        ActiveWindow = FullWindow.iloc[:,2:(4+1)]
        #print(ActiveWindow)
        Mult = 2
        Window2Pay = PlayWindow(ActiveWindow,Mult)
        WindowPay += Window2Pay
        #print(WindowPay)
        
        if Window2Pay > 0 :
            ActiveWindow = FullWindow.iloc[:,4:(6+1)]
            #print(ActiveWindow)
            Mult = 3
            Window3Pay = PlayWindow(ActiveWindow,Mult)
            WindowPay += Window3Pay
            #print(WindowPay)
            if Window3Pay > 0 :
                ActiveWindow = FullWindow.iloc[:,6:(8+1)]
                #print(ActiveWindow)
                Mult = 5
                Window4Pay = PlayWindow(ActiveWindow,Mult)
                WindowPay += Window4Pay
                #print(WindowPay)
                if Window4Pay > 0:
                    WindowPay += GetJackpotPay()
                    #print(WindowPay)
        



    
    #3ventually need to add scatter pay to the free spin win    
    #ScatterCount = GetScatterCount(FullWindow)
    
    #print(FullWindow)
    #return(str(WindowPay) + "," + str(ScatterCount))
    #print(FullWindow)
    return(WindowPay)

#f = open('results.txt','w')

def PlayGame(its):
    f = open('results.txt', 'w')
    GameWin = 0
    for i in range(its):
        GameWin += PlayBaseOnce()
        if (i+1)%100 == 0:
            print(str(float(GameWin)/Bet/i) + " " + str(i))
            f.write('%s\n' % str(float(GameWin)/Bet/i))
            #print(str(i) + " " + str('{0:.10g}'.format(GameWin/Bet/i)))
            #f.write('%s\n'  % str('{0:.10g}'.format(GameWin/Bet/i)))
            #print(GameWin/Bet/its)   
            
    f.write(str(float(GameWin)/Bet/i))        
    f.close()    
    return(float(GameWin)/Bet/i)
    #return('{0:.10g}'.format(GameWin/Bet/i))

#f.close()


    
    


#f = open('results2.txt','w')
#
#for i in range(10):
#    f.write('%s\n' % i)
#    
#f.close()
    


#0.18132619047619047
